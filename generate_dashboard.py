"""
generate_dashboard.py
Generates mechanisms_dashboard.xlsx — a 12-week mechanisms planning dashboard
with Power Query placeholders for SharePoint data and formula-driven calculations.

Run: python generate_dashboard.py
Output: mechanisms_dashboard.xlsx in the same folder
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Colours ──────────────────────────────────────────────────────────────────
C_HEADER_DARK   = "1F3864"   # dark navy – title bar
C_HEADER_MID    = "2E75B6"   # mid blue – column headers
C_WEEK_HDR      = "BDD7EE"   # pale blue – week header row
C_RED_FILL      = "FF0000"   # stockout (proj stock < 0)
C_AMBER_FILL    = "FFD966"   # below min level
C_GREEN_FILL    = "A9D18E"   # healthy stock
C_HELPER_FILL   = "F2F2F2"   # hidden helper columns
C_ALT_ROW       = "EEF4FB"   # alternate row shading
C_BORDER        = "9DC3E6"

# ── Layout constants ──────────────────────────────────────────────────────────
NUM_WEEKS       = 12
NUM_MECH_ROWS   = 100        # formula rows pre-built
DATA_START_ROW  = 5          # first mechanism row
WEEK_DATE_ROW   = 4          # row holding the Monday dates for each week

# Column indices (1-based)
COL_CODE        = 1   # A
COL_LEADTIME    = 2   # B
COL_STOCK       = 3   # C
COL_MINLEVEL    = 4   # D
COL_PRICE       = 5   # E
COL_ORDER_FLAG  = 6   # F
COL_ORDER_QTY   = 7   # G
COL_STOCK_WK1   = 8   # H  (H..S = 12 projected-stock cols)
COL_DEMAND_WK1  = COL_STOCK_WK1  + NUM_WEEKS   # T (T..AE = 12 demand helper cols)
COL_INCOMING_WK1= COL_DEMAND_WK1 + NUM_WEEKS   # AF (AF..AQ = 12 incoming helper cols)

def col(n):
    """Return Excel column letter for 1-based column index n."""
    return get_column_letter(n)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(color="FFFFFF", bold=True, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def normal_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

# ── Data-sheet helper ─────────────────────────────────────────────────────────
def create_data_sheet(wb, sheet_name, table_name, columns):
    """Create a hidden data sheet with column headers only (no table object).
    Power Query will create the table when loaded, naming it after the query.
    Rename each query table to match table_name after loading."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_state = "hidden"

    hdr_fill = make_fill("2E75B6")
    for ci, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = header_font()
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = 18

    # Note: intentionally no Table object created here — Power Query
    # creates the table on load. User renames it to table_name afterwards.
    return ws

# ── Dashboard sheet ───────────────────────────────────────────────────────────
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:S1")
    title = ws["A1"]
    title.value     = "Mechanisms Planning Dashboard — 12-Week View"
    title.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title.fill      = make_fill(C_HEADER_DARK)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value     = '="Data refreshed: "&TEXT(NOW(),"dd/mm/yyyy hh:mm")'
    sub.font      = normal_font(size=9, color="666666")
    sub.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("H2:S2")
    wk_sub = ws["H2"]
    wk_sub.value     = '="Weeks: "&TEXT(TODAY()-WEEKDAY(TODAY(),3),"dd/mm/yy")&" to "&TEXT(TODAY()-WEEKDAY(TODAY(),3)+83,"dd/mm/yy")'
    wk_sub.font      = normal_font(size=9, color="666666")
    wk_sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    hdr_fill = make_fill(C_HEADER_MID)
    static_headers = [
        ("A3", "Mech Code"),
        ("B3", "Lead-time (days)"),
        ("C3", "Current Stock"),
        ("D3", "Min Level"),
        ("E3", "Price"),
        ("F3", "Order Needed?"),
        ("G3", "Suggested Qty"),
    ]
    for addr, label in static_headers:
        c = ws[addr]
        c.value     = label
        c.font      = header_font()
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # Week headers in row 3 (merged over two sub-cols would need more cols;
    # for now just label each stock column)
    week_hdr_fill = make_fill(C_HEADER_MID)
    for k in range(NUM_WEEKS):
        c = ws.cell(row=3, column=COL_STOCK_WK1 + k,
                    value=f"Wk {k+1}")
        c.font      = header_font()
        c.fill      = week_hdr_fill
        c.alignment = Alignment(horizontal="center")

    # ── Row 4: Week date formulas ─────────────────────────────────────────────
    week_date_fill = make_fill(C_WEEK_HDR)
    for k in range(NUM_WEEKS):
        c = ws.cell(row=WEEK_DATE_ROW, column=COL_STOCK_WK1 + k)
        if k == 0:
            c.value = "=TODAY()-WEEKDAY(TODAY(),3)"
        else:
            prev = col(COL_STOCK_WK1 + k - 1)
            c.value = f"={prev}{WEEK_DATE_ROW}+7"
        c.number_format = "dd/mm/yy"
        c.font          = normal_font(bold=True, size=9)
        c.fill          = week_date_fill
        c.alignment     = Alignment(horizontal="center")

    # Blank row-4 cells for static columns
    for ci in range(COL_CODE, COL_STOCK_WK1):
        c = ws.cell(row=WEEK_DATE_ROW, column=ci)
        c.fill = week_date_fill

    ws.row_dimensions[4].height = 16

    # ── Rows 5+: Mechanism formula rows ──────────────────────────────────────
    for r in range(DATA_START_ROW, DATA_START_ROW + NUM_MECH_ROWS):
        row_fill = make_fill(C_ALT_ROW) if (r % 2 == 0) else None

        # A: Mechanism code (pulled from _Lookup in order)
        ws[f"A{r}"].value = f'=IFERROR(INDEX(_Lookup[NewCode],ROW()-{DATA_START_ROW-1}),"")'
        ws[f"A{r}"].font  = normal_font(bold=True)

        # B: Lead-time
        ws[f"B{r}"].value = f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},_Lookup[NewCode],_Lookup[LeadTime]),0))'
        ws[f"B{r}"].alignment = Alignment(horizontal="center")

        # C: Current stock
        ws[f"C{r}"].value = f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},_Sortly[MechCode],_Sortly[Quantity]),0))'
        ws[f"C{r}"].alignment = Alignment(horizontal="center")

        # D: Min level
        ws[f"D{r}"].value = f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},_Sortly[MechCode],_Sortly[MinLevel]),0))'
        ws[f"D{r}"].alignment = Alignment(horizontal="center")

        # E: Price
        ws[f"E{r}"].value = f'=IF($A{r}="","",IFERROR(XLOOKUP($A{r},_Sortly[MechCode],_Sortly[Price]),0))'
        ws[f"E{r}"].number_format = "£#,##0.00"
        ws[f"E{r}"].alignment = Alignment(horizontal="right")

        # F: Order flag
        stock_range = f"H{r}:S{r}"
        ws[f"F{r}"].value = f'=IF($A{r}="","",IF(MIN({stock_range})<$D{r},"⚠ YES","✓ OK"))'
        ws[f"F{r}"].alignment = Alignment(horizontal="center")

        # G: Suggested order quantity
        ws[f"G{r}"].value = f'=IF($A{r}="","",IF(MIN({stock_range})<$D{r},MAX($D{r}-MIN({stock_range}),0),""))'
        ws[f"G{r}"].alignment = Alignment(horizontal="center")

        # Helper cols: Demand (T..AE) and Incoming (AF..AQ)
        for k in range(NUM_WEEKS):
            wk_col   = col(COL_STOCK_WK1 + k)       # H,I,J... (week date reference)
            dem_col  = col(COL_DEMAND_WK1 + k)       # T,U,V...
            inc_col  = col(COL_INCOMING_WK1 + k)     # AF,AG,AH...

            # Demand: count production rows where this mech appears in Mech1 or Mech2
            ws[f"{dem_col}{r}"].value = (
                f'=IF($A{r}="",'
                f'"",'
                f'COUNTIFS(_Production[WeekDate],${wk_col}${WEEK_DATE_ROW},_Production[Mech1],$A{r})'
                f'+COUNTIFS(_Production[WeekDate],${wk_col}${WEEK_DATE_ROW},_Production[Mech2],$A{r}))'
            )
            ws[f"{dem_col}{r}"].fill = make_fill(C_HELPER_FILL)
            ws[f"{dem_col}{r}"].font = normal_font(size=9, color="999999")
            ws[f"{dem_col}{r}"].alignment = Alignment(horizontal="center")

            # Incoming: sum from _PO and _Seminar
            ws[f"{inc_col}{r}"].value = (
                f'=IF($A{r}="",'
                f'"",'
                f'SUMIFS(_PO[Quantity],_PO[MechCode],$A{r},_PO[WeekDate],${wk_col}${WEEK_DATE_ROW})'
                f'+SUMIFS(_Seminar[Quantity],_Seminar[MechCode],$A{r},_Seminar[WeekDate],${wk_col}${WEEK_DATE_ROW}))'
            )
            ws[f"{inc_col}{r}"].fill = make_fill(C_HELPER_FILL)
            ws[f"{inc_col}{r}"].font = normal_font(size=9, color="999999")
            ws[f"{inc_col}{r}"].alignment = Alignment(horizontal="center")

        # Projected stock columns (H..S)
        for k in range(NUM_WEEKS):
            stock_col   = col(COL_STOCK_WK1 + k)
            dem_col     = col(COL_DEMAND_WK1 + k)
            inc_col     = col(COL_INCOMING_WK1 + k)

            if k == 0:
                formula = f'=IF($A{r}="","",($C{r}+{inc_col}{r}-{dem_col}{r}))'
            else:
                prev_stock = col(COL_STOCK_WK1 + k - 1)
                formula = f'=IF($A{r}="","",{prev_stock}{r}+{inc_col}{r}-{dem_col}{r})'

            c = ws[f"{stock_col}{r}"]
            c.value     = formula
            c.alignment = Alignment(horizontal="center")
            c.font      = normal_font(bold=True)

        # Alternate row fill for static columns
        if row_fill:
            for ci in range(COL_CODE, COL_STOCK_WK1):
                existing = ws.cell(row=r, column=ci)
                if not existing.fill or existing.fill.fgColor.rgb == "00000000":
                    existing.fill = row_fill

    # ── Conditional formatting on projected stock range ───────────────────────
    stock_range = (f"H{DATA_START_ROW}:"
                   f"S{DATA_START_ROW + NUM_MECH_ROWS - 1}")

    # Red: stockout (< 0)
    ws.conditional_formatting.add(stock_range, FormulaRule(
        formula=[f'AND(H{DATA_START_ROW}<>"",H{DATA_START_ROW}<0)'],
        fill=make_fill(C_RED_FILL),
        font=Font(color="FFFFFF", bold=True)
    ))
    # Amber: below min level but not negative
    ws.conditional_formatting.add(stock_range, FormulaRule(
        formula=[f'AND(H{DATA_START_ROW}<>"",H{DATA_START_ROW}>=0,H{DATA_START_ROW}<$D{DATA_START_ROW})'],
        fill=make_fill(C_AMBER_FILL)
    ))
    # Green: at or above min level
    ws.conditional_formatting.add(stock_range, FormulaRule(
        formula=[f'AND(H{DATA_START_ROW}<>"",H{DATA_START_ROW}>=$D{DATA_START_ROW})'],
        fill=make_fill(C_GREEN_FILL)
    ))

    # Conditional formatting on F column (order flag)
    flag_range = f"F{DATA_START_ROW}:F{DATA_START_ROW + NUM_MECH_ROWS - 1}"
    ws.conditional_formatting.add(flag_range, FormulaRule(
        formula=[f'$F{DATA_START_ROW}="⚠ YES"'],
        fill=make_fill(C_AMBER_FILL),
        font=Font(bold=True, color="7F0000")
    ))

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "A": 16, "B": 13, "C": 13, "D": 10,
        "E": 10, "F": 13, "G": 13,
    }
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width
    for k in range(NUM_WEEKS):
        ws.column_dimensions[col(COL_STOCK_WK1 + k)].width = 9
    # Hide helper columns
    for k in range(NUM_WEEKS * 2):
        ws.column_dimensions[col(COL_DEMAND_WK1 + k)].width = 8
        ws.column_dimensions[col(COL_DEMAND_WK1 + k)].hidden = True

    # ── Freeze panes: freeze rows 1-4 and column A ────────────────────────────
    ws.freeze_panes = "B5"

    return ws

# ── Instructions sheet ────────────────────────────────────────────────────────
def build_instructions(wb):
    ws = wb.create_sheet("Instructions")

    def write(row, col_idx, value, bold=False, wrap=True, size=10, color="000000"):
        c = ws.cell(row=row, column=col_idx, value=value)
        c.font      = Font(name="Calibri", bold=bold, size=size, color=color)
        c.alignment = Alignment(wrap_text=wrap, vertical="top")
        return c

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 110

    title = ws.cell(row=1, column=1, value="SETUP INSTRUCTIONS — Power Query Connections")
    title.font = Font(name="Calibri", bold=True, size=13, color="1F3864")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 22

    lines = [
        (2,  True,  "OVERVIEW"),
        (3,  False, "This workbook uses 5 Power Query connections to pull live data from SharePoint."),
        (4,  False, "Once connected, press Data → Refresh All to update. The Dashboard recalculates automatically."),
        (5,  False, ""),
        (6,  True,  "STEP 1 — Open Power Query Editor"),
        (7,  False, "  In Excel: Data tab → Get Data → Launch Power Query Editor"),
        (8,  False, ""),
        (9,  True,  "STEP 2 — Create each query (repeat for all 5 below)"),
        (10, False, "  In PQ Editor: Home → New Source → Blank Query → Advanced Editor"),
        (11, False, "  Paste the M code below, click Done, then name the query exactly as shown."),
        (12, False, ""),
        (13, True,  "STEP 3 — Load each query to its sheet"),
        (14, False, "  Right-click query → Load To → select 'Table' and choose the matching sheet:"),
        (15, False, "    q_Lookup     → load to sheet _Lookup"),
        (16, False, "    q_Sortly     → load to sheet _Sortly"),
        (17, False, "    q_Production → load to sheet _Production"),
        (18, False, "    q_PO         → load to sheet _PO"),
        (19, False, "    q_Seminar    → load to sheet _Seminar"),
        (20, False, "  ⚠  After loading, right-click each table on the sheet and rename it to match"),
        (21, False, "     the query name (e.g. _Lookup, _Sortly etc.) if Excel has auto-named it differently."),
        (22, False, ""),
    ]

    for row, bold, text in lines:
        write(row, 1, "", bold)
        write(row, 2, text, bold, size=10 if not bold else 11,
              color="1F3864" if bold else "000000")
        ws.row_dimensions[row].height = 16 if not bold else 18

    # M code blocks
    queries = [
        ("q_Lookup", M_LOOKUP),
        ("q_Sortly", M_SORTLY),
        ("q_Production", M_PRODUCTION),
        ("q_PO", M_PO),
        ("q_Seminar", M_SEMINAR),
    ]

    current_row = 23
    code_fill = make_fill("F8F8F8")
    for qname, mcode in queries:
        # Query name header
        hdr = ws.cell(row=current_row, column=1, value=f"QUERY: {qname}")
        hdr.font = Font(name="Calibri", bold=True, size=11, color="2E75B6")
        ws.merge_cells(f"A{current_row}:B{current_row}")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # M code — each line in column B
        for line in mcode.strip().split("\n"):
            c = ws.cell(row=current_row, column=2, value=line)
            c.font      = Font(name="Courier New", size=9, color="1F3864")
            c.fill      = code_fill
            c.alignment = Alignment(wrap_text=False, vertical="top")
            ws.row_dimensions[current_row].height = 14
            current_row += 1

        current_row += 2  # spacer

    return ws

# ── Power Query M code ────────────────────────────────────────────────────────
SITE_URL  = "https://reposefurniturelimited.sharepoint.com/sites/ReposeFurniture-PlanningRepose"
SITE_BASE = SITE_URL + "/Shared Documents/Planning Repose/"

M_LOOKUP = f"""
let
    Source        = SharePoint.Files("{SITE_URL}", [ApiVersion = 15]),
    NavFile       = Table.SelectRows(Source, each
                        [Folder Path] = "{SITE_BASE}Mech Forecast/Mechanism Codes/"
                        and [Name] = "Mechanisms Lookup.xlsx"),
    Binary        = NavFile{{0}}[Content],
    Workbook      = Excel.Workbook(Binary, null, true),
    Sheet         = Table.SelectRows(Workbook, each [Name] = "Mechanisms lookup"){{0}}[Data],
    Headers       = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Selected      = Table.SelectColumns(Headers,
                        {{"New Code","Sage code","Sage Code 2",
                          "Seminar Code","Seminar Code 2","Lead-time"}}),
    Renamed       = Table.RenameColumns(Selected, {{
                        {{"New Code",      "NewCode"}},
                        {{"Sage code",     "SageCode"}},
                        {{"Sage Code 2",   "SageCode2"}},
                        {{"Seminar Code",  "SeminarCode"}},
                        {{"Seminar Code 2","SeminarCode2"}},
                        {{"Lead-time",     "LeadTime"}}}}),
    CleanTypes    = Table.TransformColumnTypes(Renamed,
                        {{{{"LeadTime", type number}}}}),
    Filtered      = Table.SelectRows(CleanTypes,
                        each [NewCode] <> null and [NewCode] <> "")
in
    Filtered
"""

M_SORTLY = f"""
let
    Source        = SharePoint.Files("{SITE_URL}", [ApiVersion = 15]),
    NavFile       = Table.SelectRows(Source, each
                        [Folder Path] = "{SITE_BASE}Mech Forecast/Sortly Reports/"
                        and [Name] = "Latest Sortly Mech Report.xlsx"),
    Binary        = NavFile{{0}}[Content],
    Workbook      = Excel.Workbook(Binary, null, true),
    Sheet         = Workbook{{0}}[Data],
    Headers       = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    Selected      = Table.SelectColumns(Headers,
                        {{"Entry Name","Quantity","Min Level","Price"}}),
    Renamed       = Table.RenameColumns(Selected, {{
                        {{"Entry Name", "MechCode"}},
                        {{"Min Level",  "MinLevel"}}}}),
    CleanTypes    = Table.TransformColumnTypes(Renamed, {{
                        {{"Quantity", type number}},
                        {{"MinLevel", type number}},
                        {{"Price",    type number}}}}),
    Filtered      = Table.SelectRows(CleanTypes,
                        each [MechCode] <> null and [MechCode] <> "")
in
    Filtered
"""

M_PRODUCTION = f"""
let
    Source        = SharePoint.Files("{SITE_URL}", [ApiVersion = 15]),
    NavFile       = Table.SelectRows(Source, each
                        [Folder Path] = "{SITE_BASE}"
                        and [Name] = "Production 2026 Dec-Nov.xlsx"),
    Binary        = NavFile{{0}}[Content],
    Workbook      = Excel.Workbook(Binary, null, true),

    // Keep only sheets whose name starts with "WK "
    WeekSheets    = Table.SelectRows(Workbook,
                        each Text.StartsWith([Name], "WK ")),

    // Helper: process one week sheet into a flat table
    ProcessSheet  = (sheetRec as record) as table =>
        let
            raw         = sheetRec[Data],
            // L2 = week commencing date — format "W/C DD/MM/YYYY <extra text>"
            // Take first 14 chars to strip trailing content, skip "W/C " prefix
            weekDate    = Date.FromText(Text.Middle(Text.Start(Text.From(raw{{1}}[Column12]), 14), 4), [Format="dd/MM/yyyy"]),
            // Row 4 is headers (index 3); data starts row 5
            withHdrs    = Table.PromoteHeaders(Table.Skip(raw, 3),
                              [PromoteAllScalars=true]),
            // Filter rows where ITEMS is populated
            filtered    = Table.SelectRows(withHdrs,
                              each [ITEMS] <> null and [ITEMS] <> ""),
            // Handle both dash variants in the column name
            m1Name      = if Table.HasColumns(filtered, "Mechanism \u2013 1")
                          then "Mechanism \u2013 1"
                          else "Mechanism - 1",
            m2Name      = if Table.HasColumns(filtered, "Mechanism \u2013 2")
                          then "Mechanism \u2013 2"
                          else "Mechanism - 2",
            mechCols    = Table.SelectColumns(filtered, {{m1Name, m2Name}}),
            renamed     = Table.RenameColumns(mechCols, {{
                              {{m1Name, "Mech1"}},
                              {{m2Name, "Mech2"}}}}),
            withDate    = Table.AddColumn(renamed, "WeekDate",
                              each weekDate, type date)
        in
            withDate,

    AllWeeks      = Table.Combine(
                        List.Transform(Table.ToRecords(WeekSheets),
                            each ProcessSheet(_))),
    TypedDate     = Table.TransformColumnTypes(AllWeeks,
                        {{{{"WeekDate", type date}}}}),

    // 12-week window starting from Monday of current week
    CurrentMonday = Date.StartOfWeek(Date.From(DateTime.LocalNow()), Day.Monday),
    EndDate       = Date.AddWeeks(CurrentMonday, 12),
    InWindow      = Table.SelectRows(TypedDate,
                        each [WeekDate] >= CurrentMonday
                        and  [WeekDate] < EndDate),

    // Keep rows where at least one mechanism is populated
    Filtered      = Table.SelectRows(InWindow,
                        each ([Mech1] <> null and [Mech1] <> "")
                        or   ([Mech2] <> null and [Mech2] <> ""))
in
    Filtered
"""

M_PO = f"""
let
    Source        = SharePoint.Files("{SITE_URL}", [ApiVersion = 15]),
    NavFile       = Table.SelectRows(Source, each
                        [Folder Path] = "{SITE_BASE}Mech Forecast/Purchase Orders/"
                        and [Name] = "PO Listing - with unit costs LATEST.xlsx"),
    Binary        = NavFile{{0}}[Content],
    Workbook      = Excel.Workbook(Binary, null, true),
    Sheet         = Workbook{{0}}[Data],
    // Headers in row 2 — skip row 1 then promote
    WithHeaders   = Table.PromoteHeaders(Table.Skip(Sheet, 1),
                        [PromoteAllScalars=true]),

    // Exclude Seminar supplier
    ExclSeminar   = Table.SelectRows(WithHeaders,
                        each [#"PurchaseOrder.AccountReference"] <> "SEMINARC"),

    Selected      = Table.SelectColumns(ExclSeminar, {{
                        "PurchaseOrder.AccountReference",
                        "PurchaseOrder.Date",
                        "PurchaseOrderItem.Description",
                        "PurchaseOrderItem.Quantity"}}),
    Renamed       = Table.RenameColumns(Selected, {{
                        {{"PurchaseOrder.AccountReference",   "Supplier"}},
                        {{"PurchaseOrder.Date",              "PODate"}},
                        {{"PurchaseOrderItem.Description",   "Description"}},
                        {{"PurchaseOrderItem.Quantity",      "Quantity"}}}}),
    CleanTypes    = Table.TransformColumnTypes(Renamed, {{
                        {{"PODate",   type date}},
                        {{"Quantity", type number}}}}),

    // Join to Lookup via SageCode
    LookupData    = q_Lookup,
    JoinSage1     = Table.NestedJoin(CleanTypes, "Description",
                        LookupData, "SageCode", "L1", JoinKind.LeftOuter),
    Exp1          = Table.ExpandTableColumn(JoinSage1, "L1",
                        {{"NewCode","LeadTime"}}, {{"NewCode1","LeadTime1"}}),
    JoinSage2     = Table.NestedJoin(Exp1, "Description",
                        LookupData, "SageCode2", "L2", JoinKind.LeftOuter),
    Exp2          = Table.ExpandTableColumn(JoinSage2, "L2",
                        {{"NewCode","LeadTime"}}, {{"NewCode2","LeadTime2"}}),

    // Use first non-null match
    WithMechCode  = Table.AddColumn(Exp2, "MechCode",
                        each if [NewCode1] <> null then [NewCode1] else [NewCode2]),
    WithLeadTime  = Table.AddColumn(WithMechCode, "LeadTime",
                        each if [LeadTime1] <> null then Number.From([LeadTime1])
                             else Number.From([LeadTime2])),

    // Drop unmatched lines
    Matched       = Table.SelectRows(WithLeadTime,
                        each [MechCode] <> null and [MechCode] <> ""),

    // Expected delivery = PO date + lead-time days
    WithExpected  = Table.AddColumn(Matched, "ExpectedDate",
                        each Date.AddDays([PODate], [LeadTime]), type date),

    // WeekDate = Monday of delivery week
    WithWeekDate  = Table.AddColumn(WithExpected, "WeekDate",
                        each Date.StartOfWeek([ExpectedDate], Day.Monday), type date),

    // Filter to 12-week window
    CurrentMonday = Date.StartOfWeek(Date.From(DateTime.LocalNow()), Day.Monday),
    EndDate       = Date.AddWeeks(CurrentMonday, 12),
    InWindow      = Table.SelectRows(WithWeekDate,
                        each [WeekDate] >= CurrentMonday
                        and  [WeekDate] < EndDate),

    Final         = Table.SelectColumns(InWindow,
                        {{"MechCode","Supplier","Quantity","WeekDate"}})
in
    Final
"""

M_SEMINAR = f"""
let
    Source        = SharePoint.Files("{SITE_URL}", [ApiVersion = 15]),
    NavFile       = Table.SelectRows(Source, each
                        [Folder Path] = "{SITE_BASE}Mech Forecast/Seminar/"
                        and [Name] = "Latest Repose Order Summary.xlsx"),
    Binary        = NavFile{{0}}[Content],
    Workbook      = Excel.Workbook(Binary, null, true),
    Sheet         = Table.SelectRows(Workbook,
                        each [Name] = "Seminar open orders"){{0}}[Data],
    Headers       = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),

    // Column F = Description, G = Quantity, H = Due Date (Friday of delivery week)
    Selected      = Table.SelectColumns(Headers,
                        {{"Description","Quantity","Due Date"}}),
    CleanTypes    = Table.TransformColumnTypes(Selected, {{
                        {{"Quantity",  type number}},
                        {{"Due Date",  type date}}}}),

    // ── Flatten lookup: both code columns become one "SeminarCode" column ──
    LookupData    = q_Lookup,
    LookupPart1   = Table.SelectColumns(LookupData, {{"SeminarCode", "NewCode"}}),
    LookupPart2   = Table.RenameColumns(
                        Table.SelectColumns(LookupData, {{"SeminarCode2", "NewCode"}}),
                        {{"SeminarCode2", "SeminarCode"}}),
    LookupFlat    = Table.SelectRows(
                        Table.Combine({{LookupPart1, LookupPart2}}),
                        each [SeminarCode] <> null and [SeminarCode] <> ""),

    // ── Single join – no more double-match duplication ──
    Joined        = Table.NestedJoin(CleanTypes, "Description",
                        LookupFlat, "SeminarCode", "L", JoinKind.LeftOuter),
    Expanded      = Table.ExpandTableColumn(Joined, "L",
                        {{"NewCode"}}, {{"MechCode"}}),

    // Drop unmatched
    Matched       = Table.SelectRows(Expanded,
                        each [MechCode] <> null and [MechCode] <> ""),

    // WeekDate = Monday of the week containing the due date (Friday)
    WithWeekDate  = Table.AddColumn(Matched, "WeekDate",
                        each Date.StartOfWeek([Due Date], Day.Monday), type date),

    // Filter to 12-week window
    CurrentMonday = Date.StartOfWeek(Date.From(DateTime.LocalNow()), Day.Monday),
    EndDate       = Date.AddWeeks(CurrentMonday, 12),
    InWindow      = Table.SelectRows(WithWeekDate,
                        each [WeekDate] >= CurrentMonday
                        and  [WeekDate] < EndDate),

    Final         = Table.SelectColumns(InWindow,
                        {{"MechCode","Quantity","WeekDate"}})
in
    Final
"""

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build sheets
    build_dashboard(wb)

    create_data_sheet(wb, "_Lookup",     "_Lookup",
        ["NewCode","SageCode","SageCode2","SeminarCode","SeminarCode2","LeadTime"])
    create_data_sheet(wb, "_Sortly",     "_Sortly",
        ["MechCode","Quantity","MinLevel","Price"])
    create_data_sheet(wb, "_Production", "_Production",
        ["WeekDate","Mech1","Mech2"])
    create_data_sheet(wb, "_PO",         "_PO",
        ["MechCode","Supplier","Quantity","WeekDate"])
    create_data_sheet(wb, "_Seminar",    "_Seminar",
        ["MechCode","Quantity","WeekDate"])

    build_instructions(wb)

    out_path = "mechanisms_dashboard.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"  Dashboard rows pre-built for {NUM_MECH_ROWS} mechanisms (rows {DATA_START_ROW}-{DATA_START_ROW+NUM_MECH_ROWS-1})")
    print(f"  Open Instructions sheet for Power Query setup steps.")

if __name__ == "__main__":
    main()
